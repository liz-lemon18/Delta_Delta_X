# pip install openpyxl
import openpyxl

output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "Delta X"

output_sheet["A1"] = "Name"
output_sheet["B1"] = "Delta X"

index = 2
def put_name_and_data_into_excel(name, data):
    global index
    output_sheet["A" + str(index)] = name
    output_sheet["B" + str(index)] = data
    index += 1
    
def interpolate(X, Y, target):
    output = []
    i = 0
    while i + 1 < len(Y):
        y1 = Y[i]
        y2 = Y[i + 1]        
        if (y1 <= target <= y2) or (y1 >= target >= y2):
            percent = (target - y1) / (y2 - y1)
            dx = X[i + 1] - X[i]
            output.append(X[i] + dx * percent)
        i += 1
    return output

def column_to_row(input):
    output = []
    for row in input:
        for cell in row:
            output.append(cell.value)
    return output

def iterate_excel_sheets(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        for index, sheet_name in enumerate(sheet_names):
            if index != 0:
                sheet = workbook[sheet_name]
                if type(sheet["I37"].value) is not str :                    
                    X = sheet["H37":"H49"]
                    # T1
                    T1 = sheet["I37":"I49"]
                    t1_values = interpolate(column_to_row(X), column_to_row(T1), 0.5)
                    # T4
                    T4 = sheet["K37":"K49"]
                    t4_values = interpolate(column_to_row(X), column_to_row(T4), 0.5)
                    # DELTAS
                    t1d = t1_values[1] - t1_values[0]
                    t4d = t4_values[1] - t4_values[0]
                    dd = t4d - t1d
                    put_name_and_data_into_excel(sheet_name, dd)
                else:
                    put_name_and_data_into_excel(sheet_name, "")
        output_workbook.save("output.xlsx")
    except Exception as e:
        print(f"An error occurred: {e}")

iterate_excel_sheets("7323_HbT_Healthy_Analysis_NEW.xlsx")
